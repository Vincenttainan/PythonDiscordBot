#載入函式庫=======================================================================================================================================================================================
import discord
import random 
import time
import inspect
import openpyxl
import os
from discord.ext import commands, tasks
from collections import defaultdict

#隨機在a,b間(含a,b)取得一數，回傳一數===============================================================================================================================================================
def get_rand(a, b):
	random.seed(time.time())
	return random.randint(a, b)

#取得sheet內所有資料，回傳一陣列====================================================================================================================================================================
def get_values(sheet):
	arr = []
	for row in sheet:
		arr2 = []
		for column in row:
			arr2.append(column.value)
		arr.append(arr2)
	return arr

#機器人設定，權限、聆聽前綴=========================================================================================================================================================================
intents = discord.Intents.all()
bot = commands.Bot(command_prefix="!030 ", intents=intents)

#embed顏色設定====================================================================================================================================================================================
red = 0xff0000
blue = 0x0000ff
green = 0x00ff00

#所有物品名稱陣列、所有資訊名稱陣列==================================================================================================================================================================
total_items = [
	"name", 
	"stone", 
	"coal", 
	"iron", 
	"gold", 
	"agate", 
	"wood", 
	"bamboo", 
	"old_wood", 
	"wheat", 
	"potato", 
	"carrot", 
	"gold_carrot", 
	"bread", 
	"stick", 
	"stone_dagger", 
	"stone_sword", 
	"physalis", 
	"iron_dagger", 
	"iron_sword",
	"stone_spear",
	"iron_spear"
]
total_items_chinese = [
	"名稱", 
	"石頭", 
	"煤炭", 
	"鐵礦", 
	"黃金", 
	"瑪瑙", 
	"木頭", 
	"竹子", 
	"神木", 
	"小麥", 
	"馬鈴薯", 
	"紅蘿蔔", 
	"黃金蘿蔔", 
	"麵包", 
	"木棍", 
	"石製匕首", 
	"石製長劍", 
	"酸漿", 
	"鐵製匕首", 
	"鐵製長劍",
	"石製長矛",
	"鐵製長毛"
]
total_item_dict = {
	"stone": "石頭",
	"coal": "煤炭",
	"iron": "鐵礦",
	"gold": "黃金",
	"agate": "瑪瑙",
	"wood": "木頭",
	"bamboo": "竹子",
	"old_wood": "神木",
	"wheat": "小麥",
	"potato": "馬鈴薯",
	"carrot": "紅蘿蔔",
	"gold_carrot": "黃金蘿蔔",
	"bread": "麵包",
	"stick": "木棍",
	"stone_dagger": "石製匕首",
	"stone_sword": "石製長劍",
	"physalis": "酸漿",
	"iron_dagger": "鐵製匕首",
	"iron_sword": "鐵製長劍",
	"stone_spear": "石製長矛",
	"iron_spear": "鐵製長毛"
}
user_resources = defaultdict(lambda: {
	"stone": 0, 
	"coal": 0, 
	"iron": 0, 
	"gold": 0, 
	"agate": 0, 
	"wood": 0, 
	"bamboo": 0, 
	"old_wood": 0, 
	"wheat": 0, 
	"potato": 0, 
	"carrot": 0, 
	"gold_carrot": 0, 
	"bread": 0, 
	"stick": 0, 
	"stone_dagger": 0, 
	"stone_sword": 0, 
	"physalis": 0, 
	"iron_dagger": 0, 
	"iron_sword": 0,
	"stone_spear": 0,
	"iron_spear": 0
})
total_info = [
	"name", 
	"power", 
	"money"
]
total_info_chinese = [
	"名稱", 
	"體力", 
	"錢"
]
user_info = defaultdict(lambda: {
	"power": 20, 
	"money": 0
})
total_items_attack = {
	"None": 0, 
	"stone": 1, 
	"coal": 1, 
	"iron": 1, 
	"gold": 1, 
	"agate": 1, 
	"wood": 1, 
	"bamboo": 1, 
	"old_wood": 1, 
	"wheat": 1, 
	"potato": 1, 
	"carrot": 1, 
	"gold_carrot": 1,
	"bread": 1, 
	"stick": 2, 
	"stone_dagger": 3, 
	"stone_sword": 4, 
	"physalis": 1, 
	"iron_dagger": 6, 
	"iron_sword": 7,
	"stone_spear": 3,
	"iron_spear": 6
}
mob_hp = {
	"slime": 20
}
user_shop = defaultdict(lambda: {})

#讀取excel內資訊==================================================================================================================================================================================
os.chdir(r"/Users/vincenttainan/Desktop/pythonDCbot")
wb_resources = openpyxl.load_workbook('user_resources.xlsx', data_only=True)
sheet_resources = wb_resources['工作表1']
#讀取物資excel表
_values_resources = get_values(sheet_resources)
for i in range(1, sheet_resources.max_row, 1):
	for j in range(1, len(total_items), 1):
		user_resources[int(_values_resources[i][0])][total_items[j]] = _values_resources[i][j]
#讀取資訊excel表
wb_info = openpyxl.load_workbook('user_info.xlsx', data_only=True)
sheet_info = wb_info['工作表1']
_values_info = get_values(sheet_info)
for i in range(1, sheet_info.max_row, 1):
	user_id = int(_values_info[i][0])
	for j in range(1,len(total_info),1):
		user_info[int(_values_info[i][0])][total_info[j]] = _values_info[i][j]

wb_shop = openpyxl.load_workbook('user_shop.xlsx', data_only=True)
sheet_shop = wb_shop['工作表1']
for row in sheet_shop.iter_rows(values_only=True):
	if row[0] is not None: 
		user_id = int(row[0])
		for item_info in row[1:]:
			if item_info is not None:
				item, price_count = item_info.split(': ')
				price, count = price_count.split('/')
				user_shop[user_id][item] = [int(price), int(count)]

#把bot叫起來======================================================================================================================================================================================
@bot.event
async def on_ready():
	print(f'Logged in as {bot.user.name}')
	update_energy.start()
	auto_save.start()

#回復體力=========================================================================================================================================================================================
@tasks.loop(minutes=10)
async def update_energy():
	for user_id in user_info:
		if user_info[user_id]["power"] < 20:
			user_info[user_id]["power"] += 1

#自動存檔=========================================================================================================================================================================================
@tasks.loop(minutes=60)
async def auto_save():
	sheet_resources = wb_resources['工作表1']
	for i in range(0, len(total_items), 1):
		sheet_resources.cell(row=1, column=i+1).value = total_items[i]

	i = 2
	for key, value in user_resources.items():
		sheet_resources.cell(row=i, column=1).value = str(key)
		for j in range(1, len(total_items), 1):
			sheet_resources.cell(row=i, column=j+1).value = value[total_items[j]]
		i += 1
	wb_resources.save('user_resources.xlsx')

	#儲存user_info.xlsx
	sheet_info = wb_info['工作表1']
	for i in range(0, len(total_info), 1):
		sheet_info.cell(row=1, column=i+1).value = total_info[i]

	i = 2
	for info_key, info_value in user_info.items():
		sheet_info.cell(row=i, column=1).value = str(info_key)
		for j in range(1, len(total_info), 1):
			sheet_info.cell(row=i, column=j+1).value = info_value[total_info[j]]
		i += 1
	wb_info.save('user_info.xlsx')

	sheet_shop = wb_shop['工作表1']
	sheet_shop.delete_rows(1, sheet_shop.max_row)
	for row, (user_id, shop_data) in enumerate(user_shop.items(), start=1):
		sheet_shop.cell(row=row, column=1, value=str(user_id))
		for col, (item, info) in enumerate(shop_data.items(), start=2):
			sheet_shop.cell(row=row, column=col, value=f"{item}: {info[0]}/{info[1]}")
	wb_shop.save('user_shop.xlsx')
	print("auto_saved")

#add_money指令====================================================================================================================================================================================
@bot.command()
async def add_money(ctx):
	if ctx.author.id == 830075334451003422:
		for info_key, info_value in user_info.items():
			user_info[info_key]["money"]+=100
		embed = discord.Embed(title="發錢錢$$", color=green)
		embed.add_field(name="", value="發錢完畢", inline=False)
		await ctx.send(embed=embed)
	else:
		embed = discord.Embed(title="你是誰？？？？", color=red)
		embed.add_field(name="", value="不是你是誰？？？你憑什麼叫我發錢錢？？？", inline=False)
		await ctx.send(embed=embed)

#save指令=========================================================================================================================================================================================
@bot.command()
async def save(ctx):
	if ctx.author.id == 830075334451003422:
		sheet_resources = wb_resources['工作表1']
		for i in range(0, len(total_items), 1):
			sheet_resources.cell(row=1, column=i+1).value = total_items[i]

		i = 2
		for key, value in user_resources.items():
			sheet_resources.cell(row=i, column=1).value = str(key)
			for j in range(1, len(total_items), 1):
				sheet_resources.cell(row=i, column=j+1).value = value[total_items[j]]
			i += 1
		wb_resources.save('user_resources.xlsx')

		#儲存user_info.xlsx
		sheet_info = wb_info['工作表1']
		for i in range(0, len(total_info), 1):
			sheet_info.cell(row=1, column=i+1).value = total_info[i]

		i = 2
		for info_key, info_value in user_info.items():
			sheet_info.cell(row=i, column=1).value = str(info_key)
			for j in range(1, len(total_info), 1):
				sheet_info.cell(row=i, column=j+1).value = info_value[total_info[j]]
			i += 1
		wb_info.save('user_info.xlsx')

		sheet_shop = wb_shop['工作表1']
		sheet_shop.delete_rows(1, sheet_shop.max_row)
		for row, (user_id, shop_data) in enumerate(user_shop.items(), start=1):
			sheet_shop.cell(row=row, column=1, value=str(user_id))
			for col, (item, info) in enumerate(shop_data.items(), start=2):
				sheet_shop.cell(row=row, column=col, value=f"{item}: {info[0]}/{info[1]}")
		wb_shop.save('user_shop.xlsx')

		embed = discord.Embed(title="存檔ing", color=green)
		embed.add_field(name="", value="存檔完畢", inline=False)
		await ctx.send(embed=embed)

	else:
		embed = discord.Embed(title="你是誰？？？？", color=red)
		embed.add_field(name="", value="不是你是誰？？？你憑什麼叫我存檔？？？", inline=False)
		await ctx.send(embed=embed)

#mine指令========================================================================================================================================================================================
@bot.command()
async def mine(ctx):
	user_id,user_name=ctx.author.id,ctx.author.name
	if user_info[user_id]["power"] > 0:
		user_info[user_id]["power"] -= 1
		rnd=get_rand(1,99)
		if rnd>=0 and rnd<50:
			cnt=get_rand(4,5)
			user_resources[user_id]["stone"] += cnt
			embed = discord.Embed(title="石頭", color=green)
			embed.add_field(name="", value="`"+user_name+"`"+"在礦洞口撿了`"+str(cnt)+"`顆`石頭`", inline=False)
		elif rnd>=50 and rnd<80:
			cnt=get_rand(2,4)
			user_resources[user_id]["coal"] += cnt
			embed = discord.Embed(title="煤炭", color=green)
			embed.add_field(name="", value="稍稍探索了一下，"+"`"+user_name+"`"+"發現了`"+str(cnt)+"`塊`煤炭`", inline=False)
		elif rnd>=80 and rnd<90:
			cnt=get_rand(1,3)
			user_resources[user_id]["iron"] += cnt
			embed = discord.Embed(title="鐵礦", color=green)
			embed.add_field(name="", value="`"+user_name+"`"+"發現了一條廢棄的礦道，在中找到了`"+str(cnt)+"`塊`鐵`", inline=False)
		elif rnd>=90 and rnd<96:
			cnt=get_rand(1,2)
			user_resources[user_id]["gold"] += cnt
			embed = discord.Embed(title="黃金", color=green)
			embed.add_field(name="", value="亮亮的、金光閃閃，"+"`"+user_name+"`"+"找到了`"+str(cnt)+"`塊`黃金`", inline=False)
		elif rnd>=96 and rnd<100:
			user_resources[user_id]["agate"] += 1
			embed = discord.Embed(title="瑪瑙", color=green)
			embed.add_field(name="", value="經過許久的探索，"+"`"+user_name+"`"+"終於挖到了`1`片`瑪瑙`", inline=False)
		embed.add_field(name="", value="剩餘體力`"+str(user_info[user_id]["power"])+"`",inline=False)
		await ctx.send(embed=embed)
	else:
		embed = discord.Embed(title="體力不足", color=red)
		embed.add_field(name="", value="`"+user_name+"`現在太累了，需要休息", inline=False)
		await ctx.send(embed=embed)

#wood指令========================================================================================================================================================================================
@bot.command()
async def wood(ctx):
	user_id, user_name = ctx.author.id, ctx.author.name
	if user_info[user_id]["power"] > 0:
		user_info[user_id]["power"] -= 1
		rnd = get_rand(1, 99)
		if rnd >= 0 and rnd < 70:
			cnt = get_rand(1, 4)
			user_resources[user_id]["wood"] += cnt
			embed = discord.Embed(title="木頭", color=green)
			embed.add_field(name="", value="`"+user_name+"`"+"在森林中砍了`"+str(cnt)+"`塊`木頭`", inline=False)
		elif rnd >= 70 and rnd < 95:
			cnt = get_rand(1, 3)
			user_resources[user_id]["bamboo"] += cnt
			embed = discord.Embed(title="竹子", color=green)
			embed.add_field(name="", value="`"+user_name+"`"+"找到了一片竹林，砍了`"+str(cnt)+"`塊`竹子`", inline=False)
		elif rnd >= 95 and rnd < 100:
			user_resources[user_id]["old_wood"] += 1
			embed = discord.Embed(title="神木", color=green)
			embed.add_field(name="", value="`"+user_name+"`"+"發現了一棵神奇的大樹，在它身上砍下了`1`塊`神木`", inline=False)
		embed.add_field(name="", value="剩餘體力`"+str(user_info[user_id]["power"])+"`",inline=False)
		await ctx.send(embed=embed)
	else:
		embed = discord.Embed(title="體力不足", color=red)
		embed.add_field(name="", value="`"+user_name+"`現在太累了，需要休息", inline=False)
		await ctx.send(embed=embed)

#farm指令========================================================================================================================================================================================
@bot.command()
async def farm(ctx):
	user_id, user_name = ctx.author.id, ctx.author.name
	if user_info[user_id]["power"] > 0:
		user_info[user_id]["power"] -= 1
		rnd=get_rand(1,99)
		if rnd >= 0 and rnd < 33:
			cnt = get_rand(1, 3)
			user_resources[user_id]["wheat"] += cnt
			embed = discord.Embed(title="小麥", color=green)
			embed.add_field(name="", value="`"+user_name+"`"+"經過一天的耕作，採收了`"+str(cnt)+"`個`小麥`", inline=False)
		elif rnd >= 33 and rnd < 66:
			cnt = get_rand(1, 2)
			user_resources[user_id]["potato"] += cnt
			embed = discord.Embed(title="馬鈴薯", color=green)
			embed.add_field(name="", value="`"+user_name+"`"+"翻了一天的地，找到了`"+str(cnt)+"`塊`馬鈴薯`", inline=False)
		elif rnd >= 66 and rnd < 99:
			cnt = get_rand(1, 2)
			user_resources[user_id]["carrot"] += cnt
			embed = discord.Embed(title="紅蘿蔔", color=green)
			embed.add_field(name="", value="`"+user_name+"`"+"跟著兔子的蹤跡，找到了`"+str(cnt)+"`根`紅蘿蔔`", inline=False)
		elif rnd ==99:
			user_resources[user_id]["gold_carrot"] += 1
			embed = discord.Embed(title="黃金蘿蔔", color=green)
			embed.add_field(name="", value="金光閃閃，"+"`"+user_name+"`"+"在一片神奇的田裡，拔出了`1`根`黃金蘿蔔`", inline=False)
		embed.add_field(name="", value="剩餘體力`"+str(user_info[user_id]["power"])+"`",inline=False)
		await ctx.send(embed=embed)
	else:
		embed = discord.Embed(title="體力不足", color=red)
		embed.add_field(name="", value="`"+user_name+"`現在太累了，需要休息", inline=False)
		await ctx.send(embed=embed)

#bag指令=========================================================================================================================================================================================
@bot.command()
async def bag(ctx):
	user_id, user_name = ctx.author.id, ctx.author.name
	embed = discord.Embed(title="**"+user_name+"**的包包", color=blue)
	for i in range(1,len(total_items),1):
		if user_resources[user_id][total_items[i]] != 0:
			embed.add_field(name="", value="**"+total_items_chinese[i]+"**\t(`"+str(total_items[i])+"`) : `"+str(user_resources[user_id][total_items[i]])+"` 個",inline=False)
	await ctx.send(embed=embed)

#sleep指令=======================================================================================================================================================================================
@bot.command()
async def sleep(ctx):
	if ctx.author.id == 830075334451003422:
		#儲存user_resources.xlsx
		sheet_resources = wb_resources['工作表1']
		for i in range(0, len(total_items), 1):
			sheet_resources.cell(row=1, column=i+1).value = total_items[i]

		i = 2
		for key, value in user_resources.items():
			sheet_resources.cell(row=i, column=1).value = str(key)
			for j in range(1, len(total_items), 1):
				sheet_resources.cell(row=i, column=j+1).value = value[total_items[j]]
			i += 1
		wb_resources.save('user_resources.xlsx')

		#儲存user_info.xlsx
		sheet_info = wb_info['工作表1']
		for i in range(0, len(total_info), 1):
			sheet_info.cell(row=1, column=i+1).value = total_info[i]

		i = 2
		for info_key, info_value in user_info.items():
			sheet_info.cell(row=i, column=1).value = str(info_key)
			for j in range(1, len(total_info), 1):
				sheet_info.cell(row=i, column=j+1).value = info_value[total_info[j]]
			i += 1
		wb_info.save('user_info.xlsx')

		sheet_shop = wb_shop['工作表1']
		for row, (user_id, shop_data) in enumerate(user_shop.items(), start=1):
			sheet_shop.cell(row=row, column=1, value=str(user_id))
			for col, (item, info) in enumerate(shop_data.items(), start=2):
				sheet_shop.cell(row=row, column=col, value=f"{item}: {info[0]}/{info[1]}")
		wb_shop.save('user_shop.xlsx')

		embed = discord.Embed(title="正在關機030", color=green)
		embed.add_field(name="", value="晚安，馬卡巴卡", inline=False)
		await ctx.send(embed=embed)
		await bot.close()
	else:
		embed = discord.Embed(title="你是誰？？？？", color=red)
		embed.add_field(name="", value="不是你是誰？？？你憑什麼叫我關機？？？", inline=False)
		await ctx.send(embed=embed)

#sos指令=========================================================================================================================================================================================
@bot.command()
async def sos(ctx):
	embed = discord.Embed(title="`LittleVincenttainan`的指令列表", color=blue)
	embed.add_field(name="", value="指令前綴：`!030`\n", inline=False)
	embed.add_field(name="**mine**", value="挖礦囉", inline=False)
	embed.add_field(name="**wood**", value="伐木囉", inline=False)
	embed.add_field(name="**farm**", value="種田囉", inline=False)
	embed.add_field(name="**eat `物品名稱`**", value="吃東西回復體力", inline=False)
	embed.add_field(name="**bag**", value="檢視目前背包裡有的物品", inline=False)
	embed.add_field(name="**craft `物品名稱`**", value="合成物品", inline=False)
	embed.add_field(name="**info**", value="檢視目前玩家資訊", inline=False)
	embed.add_field(name="**fight `怪物名稱` `物品名稱(選填)`**", value="戰鬥與侵略", inline=False)
	embed.add_field(name="**shop `參數` `更多參數`**", value="戰鬥與侵略", inline=False)
	embed.add_field(name="**buy `@玩家名稱` `物品名稱` `數量`**", value="戰鬥與侵略", inline=False)
	embed.add_field(name="**sleep**", value="關閉機器人，需開發者權限", inline=False)
	embed.add_field(name="**add_money**", value="給所有人發100$，需開發者權限", inline=False)
	embed.add_field(name="**save**", value="儲存玩家資訊，需開發者權限", inline=False)
	embed.add_field(name="**update**", value="查看更新日誌", inline=False)
	await ctx.send(embed=embed)

#hello指令=======================================================================================================================================================================================
@bot.command()
async def hello(ctx):
	embed = discord.Embed(title="LittleVincenttainan", color=blue)
	embed.add_field(name="簡介", value="`LittleVincenttainan`是由`Vincenttainan`餵養長大的機器人\n\n但因為`Vincenttainan`不會寫程式所以`LittleVincent`有點發育遲緩", inline=False)
	await ctx.send(embed=embed)

#info指令========================================================================================================================================================================================
@bot.command()
async def info(ctx):
	user_id, user_name = ctx.author.id, ctx.author.name
	embed = discord.Embed(title="**"+user_name+"**的資訊", color=blue)
	embed.add_field(name="", value="體力：`"+str(user_info[user_id]["power"])+"`/20", inline=False)
	embed.add_field(name="", value="錢：`"+str(user_info[user_id]["money"])+"`$", inline=False)
	await ctx.send(embed=embed, ephemeral=True)

#update指令======================================================================================================================================================================================
@bot.command()
async def update(ctx):
	embed = discord.Embed(title="LittleVincenttainan更新日誌", color=blue)
	embed.add_field(name="**20240410**", value="1.\t加入 `mine` 指令\n2.\t加入 `hello` 指令\n3.\t加入 `bag` 指令\n4.\t加入 `sos` 指令", inline=False)
	embed.add_field(name="**20240411**", value="1.\t加入 `wood` 指令\n2.\t更改 `資訊儲存` 功能\n3.\t加入 `sleep` 指令", inline=False)
	embed.add_field(name="**20240412**", value="1.\t加入 `info` 指令\n2.\t加入 `farm` 指令\n3.\t加入 `體力值` 系統\n4.\t加入 `eat` 指令", inline=False)
	embed.add_field(name="**20240413**", value="1.\t加入 `自動存擋` 功能\n2.\t加入 `craft` 指令\n", inline=False)
	embed.add_field(name="**20240414**", value="1.\t加入 `save` 指令\n", inline=False)
	embed.add_field(name="**20240415**", value="1.\t加入 `fight` 指令\n2.\t加入 `money` 系統\n3.\t加入 `add_money` 指令", inline=False)
	embed.add_field(name="**20240416**", value="1.\t加入 `shop add` 指令\n2.\t加入 `shop remove` 系統", inline=False)
	embed.add_field(name="**20240417**", value="1.\t加入 `shop @玩家` 指令\n2.\t加入 `shop 儲存` 功能", inline=False)
	embed.add_field(name="**20240418**", value="1.\t修正 `shop 儲存` 功能\n2.\t加入 `buy` 指令", inline=False)
	embed.add_field(name="**20240419**", value="1.\t修正 `buy` 指令" , inline=False)
	await ctx.send(embed=embed)

#eat指令=========================================================================================================================================================================================
@bot.command()
async def eat(ctx, *, food=None):
	user_id, user_name = ctx.author.id, ctx.author.name
	if food is None:
		embed = discord.Embed(title="吃東西 : `!030 eat` `物品名稱`", color=blue)
		embed.add_field(name="", value="吃東西回復體力\n\n但你需要告訴我你要吃什麼啊030\n", inline=False)
		embed.add_field(name="可以吃的東西", value="紅蘿蔔`carrot`剩餘`"+str(user_resources[user_id]["carrot"])+"`個：回復`1`點體力", inline=False)
		embed.add_field(name="", value="黃金蘿蔔`gold_carrot`剩餘`"+str(user_resources[user_id]["gold_carrot"])+"`個：回復`20`點體力", inline=False)
		embed.add_field(name="", value="麵包`bread`剩餘`"+str(user_resources[user_id]["bread"])+"`個：回復`3`點體力", inline=False)
		await ctx.send(embed=embed)
	else:
		#胡蘿蔔
		if food == "carrot":
			if user_resources[user_id]["carrot"] >=1:
				user_resources[user_id]["carrot"] -= 1
				user_info[user_id]["power"] += 1
				embed = discord.Embed(title="吃東西", color=green)
				embed.add_field(name="", value="吃胡蘿蔔回復體力\n\n剩餘體力：`"+str(user_info[user_id]["power"])+"`\n\n剩餘胡蘿蔔`"+str(user_resources[user_id]["carrot"])+"`個", inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="吃東西", color=red)
				embed.add_field(name="", value="吃東西回復體力\n\n但你好像沒有胡蘿蔔了", inline=False)
				await ctx.send(embed=embed)
		#麵包
		elif food == "bread":
			if user_resources[user_id]["bread"] >=1:
				user_resources[user_id]["bread"] -= 1
				user_info[user_id]["power"] += 3
				embed = discord.Embed(title="吃東西", color=green)
				embed.add_field(name="", value="吃麵包回復體力\n\n剩餘體力：`"+str(user_info[user_id]["power"])+"`\n\n剩餘麵包`"+str(user_resources[user_id]["bread"])+"`個", inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="吃東西", color=red)
				embed.add_field(name="", value="吃東西回復體力\n\n但你好像沒有麵包了", inline=False)
				await ctx.send(embed=embed)
		#黃金蘿蔔
		elif food == "gold_carrot":
			if user_resources[user_id]["gold_carrot"] >=1:
				user_resources[user_id]["gold_carrot"] -= 1
				user_info[user_id]["power"] += 20
				embed = discord.Embed(title="吃東西", color=green)
				embed.add_field(name="", value="吃黃金蘿蔔回復體力\n\n剩餘體力：`"+str(user_info[user_id]["power"])+"`\n\n剩餘黃金蘿蔔`"+str(user_resources[user_id]["gold_carrot"])+"`個", inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="吃東西", color=red)
				embed.add_field(name="", value="吃東西回復體力\n\n但你好像沒有黃金蘿蔔了", inline=False)
				await ctx.send(embed=embed)
		else:
			embed = discord.Embed(title="吃東西", color=red)
			embed.add_field(name="", value="吃東西回復體力\n\n你這東西可以吃嗎030", inline=False)
			await ctx.send(embed=embed)

#craft指令=======================================================================================================================================================================================
@bot.command()
async def craft(ctx, *, target=None):
	user_id, user_name = ctx.author.id, ctx.author.name
	if target is None:
		embed = discord.Embed(title="合成物品 : `!030 craft` `物品名稱`", color=blue)
		embed.add_field(name="", value="消耗材料合成物品\n\n但你需要告訴我你要合成什麼啊030\n",inline=False)
		embed.add_field(name="可合成的物品",value="麵包`bread`：小麥`"+str(user_resources[user_id]["wheat"])+"`/3個",inline=False)
		embed.add_field(name="",value="木棍`stick`：木頭`"+str(user_resources[user_id]["wood"])+"`/2個",inline=False)
		embed.add_field(name="",value="石製匕首`stone_dagger`：石頭`"+str(user_resources[user_id]["stone"])+"`/2個\t木棒`"+str(user_resources[user_id]["stick"])+"`/1個",inline=False)
		embed.add_field(name="",value="石製長劍`stone_sword`：石頭`"+str(user_resources[user_id]["stone"])+"`/5個\t木棒`"+str(user_resources[user_id]["stick"])+"`/2個",inline=False)
		embed.add_field(name="",value="石製長矛`stone_spear`：石頭`"+str(user_resources[user_id]["stone"])+"`/1個\t竹子`"+str(user_resources[user_id]["bamboo"])+"`/4個",inline=False)
		embed.add_field(name="",value="鐵製匕首`iron_dagger`：鐵礦`"+str(user_resources[user_id]["iron"])+"`/2個\t木棒`"+str(user_resources[user_id]["stick"])+"`/1個",inline=False)
		embed.add_field(name="",value="鐵製長劍`iron_sword`：鐵礦`"+str(user_resources[user_id]["iron"])+"`/5個\t木棒`"+str(user_resources[user_id]["stick"])+"`/2個",inline=False)
		embed.add_field(name="",value="鐵製長矛`iron_spear`：鐵礦`"+str(user_resources[user_id]["iron"])+"`/1個\t竹子`"+str(user_resources[user_id]["bamboo"])+"`/4個",inline=False)
		await ctx.send(embed=embed)
	else:
		if target == "bread":
		#麵包
			if user_resources[user_id]["wheat"] >= 3:
				user_resources[user_id]["wheat"] -= 3
				user_resources[user_id]["bread"] += 1
				embed = discord.Embed(title="合成物品", color=green)
				embed.add_field(name="", value="合成物品成功\n\n你現在有`"+str(user_resources[user_id]["bread"])+"`個麵包\n剩餘`"+str(user_resources[user_id]["wheat"])+"`個小麥",inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="合成物品", color=red)
				embed.add_field(name="",value="合成`麵包`失敗\n\n你的物品好像不夠合成", inline=False)
				embed.add_field(name="",value="麵包`bread`：小麥`"+str(user_resources[user_id]["wheat"])+"`/3個",inline=False)
				await ctx.send(embed=embed)
		elif target == "stick":
		#木棍
			if user_resources[user_id]["wood"] >= 2:
				user_resources[user_id]["wood"] -= 2
				user_resources[user_id]["stick"] += 1
				embed = discord.Embed(title="合成物品", color=green)
				embed.add_field(name="", value="合成物品成功\n\n你現在有`"+str(user_resources[user_id]["stick"])+"`個木棍\n剩餘`"+str(user_resources[user_id]["wood"])+"`個木頭",inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="合成物品", color=red)
				embed.add_field(name="",value="合成`木棍`失敗\n\n你的物品好像不夠合成", inline=False)
				embed.add_field(name="",value="木棍`stick`：木頭`"+str(user_resources[user_id]["wood"])+"`/2個",inline=False)
				await ctx.send(embed=embed)
		elif target == "stone_dagger":
		#石製匕首
			if user_resources[user_id]["stone"] >= 2 and user_resources[user_id]["stick"] >= 1:
				user_resources[user_id]["stone"] -= 2
				user_resources[user_id]["stick"] -= 1
				user_resources[user_id]["stone_dagger"] += 1
				embed = discord.Embed(title="合成物品", color=green)
				embed.add_field(name="", value="合成物品成功\n\n你現在有`"+str(user_resources[user_id]["stone_dagger"])+"`個石製匕首\n剩餘`"+str(user_resources[user_id]["stone"])+"`個石頭\t`"+str(user_resources[user_id]["stick"])+"`個木棒",inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="合成物品", color=red)
				embed.add_field(name="",value="合成`石製匕首`失敗\n\n你的物品好像不夠合成", inline=False)
				embed.add_field(name="",value="石製匕首`stone_dagger`：石頭`"+str(user_resources[user_id]["stone"])+"`/2個\t木棒`"+str(user_resources[user_id]["stick"])+"`/1個",inline=False)
				await ctx.send(embed=embed)
		elif target == "stone_sword":
		#石製長劍
			if user_resources[user_id]["stone"] >= 5 and user_resources[user_id]["stick"] >= 2:
				user_resources[user_id]["stone"] -= 5
				user_resources[user_id]["stick"] -= 2
				user_resources[user_id]["stone_sword"] += 1
				embed = discord.Embed(title="合成物品", color=green)
				embed.add_field(name="", value="合成物品成功\n\n你現在有`"+str(user_resources[user_id]["stone_sword"])+"`個石製長劍\n剩餘`"+str(user_resources[user_id]["stone"])+"`個石頭\t`"+str(user_resources[user_id]["stick"])+"`個木棒",inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="合成物品", color=red)
				embed.add_field(name="",value="合成`石製長劍`失敗\n\n你的物品好像不夠合成", inline=False)
				embed.add_field(name="",value="石製長劍`stone_sword`：石頭`"+str(user_resources[user_id]["stone"])+"`/5個\t木棒`"+str(user_resources[user_id]["stick"])+"`/2個",inline=False)
				await ctx.send(embed=embed)
		elif target == "iron_dagger":
		#鐵製匕首
			if user_resources[user_id]["iron"] >= 2 and user_resources[user_id]["stick"] >= 1:
				user_resources[user_id]["iron"] -= 2
				user_resources[user_id]["stick"] -= 1
				user_resources[user_id]["iron_dagger"] += 1
				embed = discord.Embed(title="合成物品", color=green)
				embed.add_field(name="", value="合成物品成功\n\n你現在有`"+str(user_resources[user_id]["iron_dagger"])+"`個鐵製匕首\n剩餘`"+str(user_resources[user_id]["iron"])+"`個鐵礦\t`"+str(user_resources[user_id]["stick"])+"`個木棒",inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="合成物品", color=red)
				embed.add_field(name="",value="合成`鐵製匕首`失敗\n\n你的物品好像不夠合成", inline=False)
				embed.add_field(name="",value="鐵製匕首`stone_dagger`：鐵礦`"+str(user_resources[user_id]["iron"])+"`/2個\t木棒`"+str(user_resources[user_id]["stick"])+"`/1個",inline=False)
				await ctx.send(embed=embed)
		elif target == "iron_sword":
		#鐵製長劍
			if user_resources[user_id]["iron"] >= 5 and user_resources[user_id]["stick"] >= 2:
				user_resources[user_id]["iron"] -= 5
				user_resources[user_id]["stick"] -= 2
				user_resources[user_id]["iron_sword"] += 1
				embed = discord.Embed(title="合成物品", color=green)
				embed.add_field(name="", value="合成物品成功\n\n你現在有`"+str(user_resources[user_id]["iron_sword"])+"`個鐵製長劍\n剩餘`"+str(user_resources[user_id]["iron"])+"`個鐵礦\t`"+str(user_resources[user_id]["stick"])+"`個木棒",inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="合成物品", color=red)
				embed.add_field(name="",value="合成`鐵製長劍`失敗\n\n你的物品好像不夠合成", inline=False)
				embed.add_field(name="",value="鐵製長劍`iron_sword`：鐵礦`"+str(user_resources[user_id]["iron"])+"`/5個\t木棒`"+str(user_resources[user_id]["stick"])+"`/2個",inline=False)
				await ctx.send(embed=embed)
		elif target == "stone_spear":
		#石製長矛
			if user_resources[user_id]["stone"] >= 1 and user_resources[user_id]["bamboo"] >= 4:
				user_resources[user_id]["stone"] -= 1
				user_resources[user_id]["bamboo"] -= 4
				user_resources[user_id]["stone_spear"] += 1
				embed = discord.Embed(title="合成物品", color=green)
				embed.add_field(name="", value="合成物品成功\n\n你現在有`"+str(user_resources[user_id]["stone_spear"])+"`個石製長矛\n剩餘`"+str(user_resources[user_id]["stone"])+"`個石頭\t`"+str(user_resources[user_id]["bamboo"])+"`個竹子",inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="合成物品", color=red)
				embed.add_field(name="",value="合成`石製長矛`失敗\n\n你的物品好像不夠合成", inline=False)
				embed.add_field(name="",value="石製長矛`stone_spear`：石頭`"+str(user_resources[user_id]["stone"])+"`/1個\t竹子`"+str(user_resources[user_id]["bamboo"])+"`/4個",inline=False)
				await ctx.send(embed=embed)
		elif target == "iron_spear":
		#鐵製長矛
			if user_resources[user_id]["iron"] >= 1 and user_resources[user_id]["bamboo"] >= 4:
				user_resources[user_id]["iron"] -= 1
				user_resources[user_id]["bamboo"] -= 4
				user_resources[user_id]["iron_spear"] += 1
				embed = discord.Embed(title="合成物品", color=green)
				embed.add_field(name="", value="合成物品成功\n\n你現在有`"+str(user_resources[user_id]["iron_spear"])+"`個鐵製長矛\n剩餘`"+str(user_resources[user_id]["iron"])+"`個鐵礦\t`"+str(user_resources[user_id]["bamboo"])+"`個竹子",inline=False)
				await ctx.send(embed=embed)
			else:
				embed = discord.Embed(title="合成物品", color=red)
				embed.add_field(name="",value="合成`鐵製長矛`失敗\n\n你的物品好像不夠合成", inline=False)
				embed.add_field(name="",value="鐵製長矛`iron_spear`：鐵礦`"+str(user_resources[user_id]["iron"])+"`/1個\t竹子`"+str(user_resources[user_id]["bamboo"])+"`/4個",inline=False)
				await ctx.send(embed=embed)
		else:
		#沒有的
			embed = discord.Embed(title="合成物品", color=red)
			embed.add_field(name="", value="消耗材料合成物品\n\n這種東西我不會合成誒030", inline=False)
			await ctx.send(embed=embed)

#fight指令=======================================================================================================================================================================================
@bot.command()
async def fight(ctx, *, inputs=None):
	user_id, user_name = ctx.author.id, ctx.author.name
	flg,attack=0,1
	if inputs is None:
		target,weapon=None,None
	elif " "in inputs:
		target, weapon=inputs.split()
	else:
		target = inputs
		weapon = None
	if target is None:
		embed = discord.Embed(title="攻擊&侵略 : `!030 fight` `怪物名稱` `物品名稱(選填)`", color=blue)
		embed.add_field(name="", value="消耗體力、裝備進行攻擊\n\n但你需要告訴我你攻擊的目標是什麼啊030\n",inline=False)
		embed.add_field(name="可攻擊的目標",value="史萊姆`slime`：\n| 目前血量：`"+str(mob_hp["slime"])+"`/20\n| 所需傷害`1`點\n| 所需體力`1`點",inline=False)
		await ctx.send(embed=embed)
		return
	if weapon is None:
		flg=1
		attack+=0
	else:
		if weapon in total_items_attack:
			if user_resources[user_id][weapon]>=1:
				user_resources[user_id][weapon]-=1
				attack+=total_items_attack[weapon]
				flg=1
			else:
				flg=0
				embed = discord.Embed(title="攻擊&侵略", color=red)
				embed.add_field(name="", value="消耗體力、裝備進行攻擊\n\n但你的好像沒有`"+weapon+"`了啊030\n",inline=False)
				await ctx.send(embed=embed)
				return
		else:
			flg=0
	if flg:
		#史萊姆
		if target == "slime":
			if user_info[user_id]["power"] > 0:
				user_info[user_id]["power"] -= 1
				embed = discord.Embed(title="攻擊&侵略", color=blue)
				embed.add_field(name="`史萊姆`出現啦", value="目前傷害：`"+str(attack)+"`點",inline=False)
				await ctx.send(embed=embed)
				mob_hp["slime"]=mob_hp["slime"]-attack
				if mob_hp["slime"]>0:
					embed = discord.Embed(title="攻擊&侵略", color=green)
					embed.add_field(name="`史萊姆`暫時退走了", value="目前血量：`"+str(mob_hp["slime"])+"`/20 點",inline=False)
					embed.add_field(name="", value="剩餘體力`"+str(user_info[user_id]["power"])+"`",inline=False)
					await ctx.send(embed=embed)
					return
				else:
					embed = discord.Embed(title="攻擊&侵略", color=green)
					embed.add_field(name="`史萊姆`死掉了", value="獲得了`5`個酸漿",inline=False)
					embed.add_field(name="", value="剩餘體力`"+str(user_info[user_id]["power"])+"`",inline=False)
					await ctx.send(embed=embed)
					mob_hp["slime"]=20
					user_resources[user_id]["physalis"]+=5
					return
			else:
				embed = discord.Embed(title="體力不足", color=red)
				embed.add_field(name="", value="`"+user_name+"`現在太累了，需要休息", inline=False)
				await ctx.send(embed=embed)
				return
		else:
			embed = discord.Embed(title="攻擊&侵略", color=red)
			embed.add_field(name="", value="消耗體力、裝備進行攻擊\n\n但這傢伙是誰啊030",inline=False)
			await ctx.send(embed=embed)
			return
	else:
		embed = discord.Embed(title="攻擊&侵略", color=red)
		embed.add_field(name="", value="消耗體力、裝備進行攻擊\n\n但你的武器是什麼啊030\n",inline=False)
		await ctx.send(embed=embed)
		return

#test指令========================================================================================================================================================================================
@bot.command()
async def test(ctx, *, target):
	user_id, user_name = ctx.author.id, ctx.author.name
	print(target)

#shop指令========================================================================================================================================================================================
@bot.command()
async def shop(ctx, *, inputs=None):
	user_id, user_name = ctx.author.id, ctx.author.name
	if inputs is None:
		embed = discord.Embed(title="交易 : \n", color=blue)
		embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n",inline=False)
		embed.add_field(name="", value="以`價格`上架`數量(選填)`個`物品名稱` : \n\n`!030 shop` `add` `物品名稱` `價格` `數量(選填)`\n\n\n移除商店上所有`物品名稱` : \n\n`!030 shop` `remove` `物品名稱`\n\n\n顯示`玩家名稱`目前的商店 : \n\n`!030 shop` `@玩家名稱`",inline=False)
		await ctx.send(embed=embed)
		return
	splited_input=inputs.split()
	if splited_input[0]=="add":
		if len(splited_input)==3:
			item,price,count=splited_input[1],splited_input[2],"1"
		elif len(splited_input)==4:
			item,price,count=splited_input[1],splited_input[2],splited_input[3]
		else:
			embed = discord.Embed(title="交易 : \n", color=red)
			embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但我看不懂你要做什麼誒030",inline=False)
			await ctx.send(embed=embed)
			return
		if price.isdigit() and count.isdigit():
			if int(price)>=1 and int(count)>=1:
				if item in total_items:
					if user_resources[user_id][item] >= int(count):
						user_resources[user_id][item]-=int(count)
						if item in user_shop[user_id]:
							user_shop[user_id][item][0]=int(price)
							user_shop[user_id][item][1]+=int(count)
						else:
							user_shop[user_id].update({item : [int(price),int(count)]})
						embed = discord.Embed(title="交易 : \n", color=green)
						embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n現在架上有 `"+str(user_shop[user_id][item][1])+"` 個 `"+total_item_dict[str(item)]+"`，售價為 `"+str(user_shop[user_id][item][0])+"$`",inline=False)
						await ctx.send(embed=embed)
					else:
						embed = discord.Embed(title="交易 : \n", color=red)
						embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n可是你好像沒有這個東西誒030",inline=False)
						await ctx.send(embed=embed)
						return
				else:
					embed = discord.Embed(title="交易 : \n", color=red)
					embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但我不認識你的商品誒030",inline=False)
					await ctx.send(embed=embed)
					return
			else:
				embed = discord.Embed(title="交易 : \n", color=red)
				embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但你的價格、數量不大於0誒030",inline=False)
				await ctx.send(embed=embed)
				return
		else:
			embed = discord.Embed(title="交易 : \n", color=red)
			embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n為什麼你的價格、數量不是數字030",inline=False)
			await ctx.send(embed=embed)
			return
	elif splited_input[0]=="remove":
		item=splited_input[1]
		if item in total_items:
			if item in user_shop[user_id]:
				user_resources[user_id][item]+=int(user_shop[user_id][item][1])
				embed = discord.Embed(title="交易 : \n", color=green)
				embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n你成功下架了 `"+str(user_shop[user_id][item][1])+"` 個 `"+total_item_dict[str(item)]+"` ",inline=False)
				await ctx.send(embed=embed)
				del user_shop[user_id][item]
				return
			else:
				embed = discord.Embed(title="交易 : \n", color=red)
				embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但你的商店沒有這個商品誒030",inline=False)
				await ctx.send(embed=embed)
				return
		else:
			embed = discord.Embed(title="交易 : \n", color=red)
			embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但我不認識你的商品誒030",inline=False)
			await ctx.send(embed=embed)
			return
	elif splited_input[0][0]=="<" and splited_input[0][1]=="@" and splited_input[0][-1]==">":
		shop_name=splited_input[0][2:-1]
		if not(user_shop.get(int(shop_name),None) is None) and len(user_shop.get(int(shop_name),{}))>=1:
			embed = discord.Embed(title="交易 : \n", color=blue)
			embed.add_field(name="", value="商店系統 這算不算一種期貨？\n",inline=False)
			shop=user_shop.get(int(shop_name))
			for key in shop:
				embed.add_field(name="",value=total_item_dict[str(key)]+"(`"+str(key)+"`) : `"+str(shop[key][0])+"$`/個，貨架上現有`"+str(shop[key][1])+"`個",inline=False)
			await ctx.send(embed=embed)
			return
		else:
			embed = discord.Embed(title="交易 : \n", color=red)
			embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但他的商店內沒有東西誒030",inline=False)
			await ctx.send(embed=embed)
			return
	else:
		embed = discord.Embed(title="交易 : \n", color=red)
		embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但我看不懂你要做什麼誒030",inline=False)
		await ctx.send(embed=embed)
		return

#test指令========================================================================================================================================================================================
@bot.command()
async def buy(ctx, *, inputs=None):
	user_id, user_name = ctx.author.id, ctx.author.name
	if inputs is None:
		embed = discord.Embed(title="購買 : `!030 buy` `@玩家名稱` `商品名稱` `購買數量`\n", color=blue)
		embed.add_field(name="", value="商店系統 這算不算一種期貨？",inline=False)
		await ctx.send(embed=embed)
		return
	else:
		splited_input=inputs.split()
		if len(splited_input) == 3:
			if splited_input[0][0]=="<" and splited_input[0][1]=="@" and splited_input[0][-1]==">":
				shop_name=splited_input[0][2:-1]
				item=splited_input[1]
				count=splited_input[2]
				if item in user_shop[int(shop_name)]:
					if count.isdigit():
						if int(count)>=1:
							if user_shop[int(shop_name)][item][1]>=int(count):
								if user_info[user_id]["money"]-user_shop[int(shop_name)][item][0]*int(count) >= 0:
									user_shop[int(shop_name)][item][1]-=int(count)
									user_resources[user_id][item]+=int(count)
									user_info[user_id]["money"]-=user_shop[int(shop_name)][item][0]*int(count)
									user_info[int(shop_name)]["money"]+=user_shop[int(shop_name)][item][0]*int(count)
									embed = discord.Embed(title="交易 : \n", color=green)
									embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n成功購買 `"+str(count)+"` 個 `"+str(total_item_dict[item])+"` ",inline=False)
									await ctx.send(embed=embed)
									if user_shop[int(shop_name)][item][1]==0:
										del user_shop[int(shop_name)][item]
									return
								else:
									embed = discord.Embed(title="交易 : \n", color=red)
									embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但你的錢好像不夠誒030",inline=False)
									await ctx.send(embed=embed)
									return
							else:
								embed = discord.Embed(title="交易 : \n", color=red)
								embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但他商店內數量不夠誒030",inline=False)
								await ctx.send(embed=embed)
								return
						else:
							embed = discord.Embed(title="交易 : \n", color=red)
							embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但你的數量不大於0誒030",inline=False)
							await ctx.send(embed=embed)
							return
					else:
						embed = discord.Embed(title="交易 : \n", color=red)
						embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n為什麼你的數量不是數字030",inline=False)
						await ctx.send(embed=embed)
						return
				else:
					embed = discord.Embed(title="交易 : \n", color=red)
					embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但他的商店內沒有這東西誒030",inline=False)
					await ctx.send(embed=embed)
					return
			else:
				embed = discord.Embed(title="交易 : \n", color=red)
				embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但我看不懂你要做什麼誒030",inline=False)
				await ctx.send(embed=embed)
				return
		else:
			embed = discord.Embed(title="購買 : \n", color=red)
			embed.add_field(name="", value="商店系統 這算不算一種期貨？\n\n但我看不懂你要做什麼誒030",inline=False)
			await ctx.send(embed=embed)

























#開機============================================================================================================================================================================================
bot.run("")
